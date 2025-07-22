# core/report_generator.py
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

def separar_nombre_completo(nombre_completo):
    """Separa nombres y apellidos"""
    partes = nombre_completo.strip().split()
    if len(partes) == 0:
        return ("", "", "")
    elif len(partes) == 1:
        return (partes[0], "", "")
    elif len(partes) == 2:
        return (partes[0], partes[1], "")
    else:
        return (" ".join(partes[:-2]), partes[-2], partes[-1])

def aplicar_formato_excel(ws):
    """Aplica formato profesional al Excel"""
    # Bordes delgados
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Estilo cabeceras (verde claro + negrita)
    header_fill = PatternFill(
        start_color='92D050',
        end_color='92D050',
        fill_type='solid'
    )
    header_font = Font(bold=True, color='000000')

    # Aplicar bordes a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formato especial para cabeceras
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajustar anchos de columna
    column_widths = {
        'A': 15,  # FECHA
        'B': 12,   # DNI
        'C': 15,   # PRIMER APELLIDO
        'D': 15,   # SEGUNDO APELLIDO
        'E': 20,   # NOMBRE
        'F': 12,   # N° LICENCIA
        'G': 25    # INSTRUCTOR
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Configurar página
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1

def generar_reporte_estudiantes(datos_pdf, estudiantes_info, output_dir):
    """Genera el Excel con formato automático"""
    # Preparar datos
    reporte_data = []
    for dni, nombre_completo in estudiantes_info:
        nombre, primer_ap, segundo_ap = separar_nombre_completo(nombre_completo)
        reporte_data.append({
            "FECHA PRACTICAS": f"{datos_pdf.get('D_DIA', '')}/{datos_pdf.get('D_MES', '')}/{datos_pdf.get('D_ANY', '')}",
            "DNI": dni,
            "PRIMER APELLIDO": primer_ap,
            "SEGUNDO APELLIDO": segundo_ap,
            "NOMBRE": nombre,
            "N° LICENCIA": "",
            "INSTRUCTOR": datos_pdf.get("A_INSTR", "")
        })

    # Crear DataFrame
    df = pd.DataFrame(reporte_data)

    # Ordenar columnas correctamente
    column_order = ["FECHA PRACTICAS", "DNI", "PRIMER APELLIDO",
                    "SEGUNDO APELLIDO", "NOMBRE", "N° LICENCIA", "INSTRUCTOR"]
    df = df[column_order]

    # Guardar Excel inicial
    reporte_path = output_dir / "Reporte_estudiantes.xlsx"
    df.to_excel(reporte_path, index=False, engine='openpyxl')

    # Aplicar formato profesional
    wb = load_workbook(reporte_path)
    ws = wb.active
    aplicar_formato_excel(ws)
    wb.save(reporte_path)

    return str(reporte_path)
